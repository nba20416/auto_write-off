#!/usr/bin/env python3
"""
智能核销系统 — 自动匹配销售登款表与银行流水表

用法:
    python reconcile.py <销售表.xlsx> <银行1.xlsx> [银行2.xlsx ...]

示例:
    python reconcile.py 销售登款表.xlsx 招商银行.xlsx
    python reconcile.py 销售登款表.xlsx 招商银行.xlsx 工商银行.xlsx 建设银行.xlsx

规则:
    - 仅按「金额」精确匹配（精确到分）
    - 匹配到 1 条 → 绿色高亮 + 标记"已完成"
    - 匹配到 0 条 → 不处理
    - 匹配到 ≥2 条 → 标记"待人工处理"，不自动核销
"""

from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 常量
# ---------------------------------------------------------------------------

# 绿色填充（已完成）
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# 表头样式
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)

# 默认配置（config.json 不存在时使用）
DEFAULT_CONFIG: dict = {
    "sales_table": {
        "amount_column": "金额",
    },
    "bank_table": {
        "amount_column": "金额",
    },
}


# ---------------------------------------------------------------------------
# 配置加载
# ---------------------------------------------------------------------------

def load_config() -> dict:
    """加载 config.json，若文件不存在则返回默认配置。"""
    config_path = Path(__file__).parent / "config.json"
    if config_path.exists():
        with open(config_path, "r", encoding="utf-8") as f:
            user_config = json.load(f)
        # 合并用户配置与默认配置
        merged = DEFAULT_CONFIG.copy()
        merged.update(user_config)
        return merged
    return DEFAULT_CONFIG


# ---------------------------------------------------------------------------
# 金额解析
# ---------------------------------------------------------------------------

def parse_amount(value) -> Decimal | None:
    """
    将 Excel 单元格值解析为 Decimal 金额（精确到分）。

    支持:
        - 纯数字 (int/float)
        - 千分位格式: "1,234.56"
        - 带货币符号: "¥1,234.56", "￥1,234.56"
        - 前后空格
        - 负数: "-1,234.56"、"(1,234.56)"

    返回 None 表示无法解析（空值、非数字文本等）。
    """
    if value is None:
        return None

    # --- 数字类型：先四舍五入到分，再转 Decimal ---
    if isinstance(value, (int, float)):
        rounded = round(float(value), 2)
        return Decimal(str(rounded)).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )

    # --- 字符串类型：清理格式后解析 ---
    raw = str(value).strip()
    if not raw:
        return None

    # 处理会计负号格式 (1,234.56) → -1,234.56
    is_negative = False
    if raw.startswith("(") and raw.endswith(")"):
        is_negative = True
        raw = raw[1:-1]

    # 移除货币符号、千分位逗号、空格
    cleaned = re.sub(r"[¥￥$,\s]", "", raw)

    # 处理可能的前导负号
    if cleaned.startswith("-"):
        is_negative = not is_negative  # toggle
        cleaned = cleaned[1:]

    if not cleaned:
        return None

    try:
        amount = Decimal(cleaned).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except Exception:
        return None

    return -amount if is_negative else amount


# ---------------------------------------------------------------------------
# 辅助函数
# ---------------------------------------------------------------------------

def find_column_index(ws, column_name: str, header_row: int = 1) -> int | None:
    """
    在表头行中查找列名，返回列号（从 1 开始）。
    找不到返回 None。
    """
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=header_row, column=col).value
        if cell_value and str(cell_value).strip() == column_name:
            return col
    return None


# ---------------------------------------------------------------------------
# 银行流水索引构建
# ---------------------------------------------------------------------------

def build_bank_index(
    bank_files: list[Path], amount_column_name: str
) -> dict[Decimal, list[tuple[str, int]]]:
    """
    读取所有银行流水文件，构建金额 → 匹配列表 的索引。

    参数:
        bank_files: 银行文件路径列表
        amount_column_name: 银行表中金额列的列名

    返回:
        { Decimal("1234.56"): [("招商银行", 5), ("工商银行", 12)], ... }
    """
    index: dict[Decimal, list[tuple[str, int]]] = defaultdict(list)

    for bank_path in bank_files:
        bank_name = bank_path.stem  # 文件名去掉 .xlsx 后缀 = 银行名

        print(f"  📂 读取银行流水: {bank_path.name} ... ", end="")

        try:
            wb = openpyxl.load_workbook(bank_path, data_only=True)
        except Exception as e:
            print(f"❌ 无法打开: {e}")
            continue

        ws = wb.active
        amount_col = find_column_index(ws, amount_column_name)

        if amount_col is None:
            print(f"❌ 找不到金额列「{amount_column_name}」")
            wb.close()
            continue

        count = 0
        for row in range(2, ws.max_row + 1):
            amount = parse_amount(ws.cell(row=row, column=amount_col).value)
            if amount is not None:
                index[amount].append((bank_name, row))
                count += 1

        wb.close()
        print(f"✓ {count} 条有效记录")

    return dict(index)


# ---------------------------------------------------------------------------
# 核销处理
# ---------------------------------------------------------------------------

def reconcile(
    sales_path: Path,
    bank_files: list[Path],
    config: dict,
) -> Path:
    """
    执行核销处理，返回输出文件路径。
    """
    sales_cfg = config.get("sales_table", {})
    bank_cfg = config.get("bank_table", {})

    sales_amount_col_name = sales_cfg.get("amount_column", "金额")
    bank_amount_col_name = bank_cfg.get("amount_column", "金额")

    # ---- 1. 构建银行流水索引 ----
    print("\n🔍 第一步：读取银行流水表，构建金额索引")
    print("-" * 50)

    amount_index = build_bank_index(bank_files, bank_amount_col_name)
    total_bank_records = sum(len(v) for v in amount_index.values())
    print(f"\n  📊 银行流水合计: {total_bank_records} 条，涉及 {len(amount_index)} 个不同金额\n")

    # ---- 2. 读取销售登款表 ----
    print("📋 第二步：逐行核销销售登款表")
    print("-" * 50)

    wb = openpyxl.load_workbook(sales_path)
    ws = wb.active

    # 查找金额列
    amount_col = find_column_index(ws, sales_amount_col_name)
    if amount_col is None:
        print(f"❌ 错误: 销售表中找不到金额列「{sales_amount_col_name}」")
        print(f"   表头内容: {[ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]}")
        wb.close()
        sys.exit(1)

    # ---- 3. 添加新列 ----
    last_col = ws.max_column
    status_col = last_col + 1   # 核销状态
    source_col = last_col + 2   # 匹配来源

    # 写表头
    ws.cell(row=1, column=status_col, value="核销状态")
    ws.cell(row=1, column=source_col, value="匹配来源")

    # 表头样式
    for c in (status_col, source_col):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # ---- 4. 逐行匹配 ----
    stats = {"completed": 0, "pending_manual": 0, "unmatched": 0}

    for row in range(2, ws.max_row + 1):
        amount = parse_amount(ws.cell(row=row, column=amount_col).value)

        if amount is None:
            # 金额为空或无法解析，跳过
            continue

        matches = amount_index.get(amount, [])  # [(bank_name, row_num), ...]
        match_count = len(matches)

        if match_count == 1:
            # ✅ 唯一匹配 → 自动核销
            bank_name, _bank_row = matches[0]

            # 绿色高亮整行
            for col in range(1, last_col + 1):
                ws.cell(row=row, column=col).fill = GREEN_FILL

            ws.cell(row=row, column=status_col, value="已完成")
            ws.cell(row=row, column=source_col, value=bank_name)

            stats["completed"] += 1

        elif match_count >= 2:
            # ⚠️ 多个匹配 → 待人工处理
            bank_names = "、".join(b for b, _ in matches)
            ws.cell(row=row, column=status_col, value="待人工处理")
            ws.cell(row=row, column=source_col, value=bank_names)

            stats["pending_manual"] += 1

        else:
            # match_count == 0 → 未匹配，不处理
            stats["unmatched"] += 1

    # ---- 5. 调整列宽 ----
    ws.column_dimensions[get_column_letter(status_col)].width = 14
    ws.column_dimensions[get_column_letter(source_col)].width = 20

    # ---- 6. 保存 ----
    output_path = sales_path.parent / f"{sales_path.stem}_已核销.xlsx"
    wb.save(output_path)
    wb.close()

    # ---- 7. 打印汇总 ----
    total = stats["completed"] + stats["pending_manual"] + stats["unmatched"]
    print(f"\n{'=' * 50}")
    print(f"  📊 核销汇总")
    print(f"{'=' * 50}")
    print(f"  总记录数:       {total:>6}")
    print(f"  已完成:         {stats['completed']:>6}  ({stats['completed'] / total * 100:5.1f}%)" if total else "")
    print(f"  待人工处理:     {stats['pending_manual']:>6}  ({stats['pending_manual'] / total * 100:5.1f}%)" if total else "")
    print(f"  未匹配:         {stats['unmatched']:>6}  ({stats['unmatched'] / total * 100:5.1f}%)" if total else "")
    print(f"{'=' * 50}")
    print(f"\n✅ 输出文件: {output_path}")

    return output_path


# ---------------------------------------------------------------------------
# 命令行入口
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 3:
        print(__doc__)
        print("❌ 至少需要 2 个参数: 销售表 和 至少 1 个银行表")
        print(f"\n示例: python {Path(__file__).name} 销售登款表.xlsx 招商银行.xlsx")
        sys.exit(1)

    sales_path = Path(sys.argv[1])
    bank_paths = [Path(p) for p in sys.argv[2:]]

    # 校验文件存在
    errors = []
    if not sales_path.exists():
        errors.append(f"  ✗ 销售登款表不存在: {sales_path}")
    if not sales_path.suffix.lower() in (".xlsx", ".xlsm"):
        errors.append(f"  ✗ 销售登款表格式不正确（需 .xlsx）: {sales_path}")

    for bp in bank_paths:
        if not bp.exists():
            errors.append(f"  ✗ 银行流水表不存在: {bp}")
        if bp.suffix.lower() not in (".xlsx", ".xlsm"):
            errors.append(f"  ✗ 银行流水表格式不正确（需 .xlsx）: {bp}")

    if errors:
        print("❌ 文件校验失败:")
        for e in errors:
            print(e)
        sys.exit(1)

    # 加载配置
    config = load_config()

    print(f"\n🚀 智能核销系统")
    print(f"  销售登款表: {sales_path.name}")
    print(f"  银行流水表: {', '.join(p.name for p in bank_paths)}")
    print(f"  金额列名: 销售表「{config['sales_table'].get('amount_column', '金额')}」 / 银行表「{config['bank_table'].get('amount_column', '金额')}」")

    # 执行核销
    reconcile(sales_path, bank_paths, config)


if __name__ == "__main__":
    main()
